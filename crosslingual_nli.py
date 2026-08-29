#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
crosslingual_nli.py — 언어 교차 NLI 로 번역 검증이 가능한가
===========================================================
가설:
    ④ 한국어 검증에 쓰는 mDeBERTa NLI 를 ⑥ 번역 검증에도 쓸 수 있다.
    premise = 한국어 원문, hypothesis = 외국어 번역문 으로 넣는 것이다.
    가능하다면 검증 엔진이 하나로 통일되고, CometKiwi 를 미채택한 뒤
    비어 있던 '참조 불필요 번역 검증기' 자리를 채울 수 있다.

주의할 점:
    mDeBERTa 의 XNLI 학습 언어에 베트남어·중국어는 있으나 한국어는 없다.
    한국어는 사전학습 단계에서만 들어갔다. 따라서 premise 가 한국어인
    교차 조건이 작동한다는 보장은 없다. 그래서 측정한다.

설계:
    AI Hub 의료 병렬 말뭉치(사람 검수 완료)를 정답으로 쓴다.
      정답쌍  (한국어 원문, 올바른 번역)      → 함의되어야 한다
      훼손쌍  (한국어 원문, 훼손된 번역)      → 함의되지 않아야 한다
    라벨을 사람이 만들 필요가 없다. 훼손 방식 자체가 라벨이다.

훼손 4종:
    MISMATCH    다른 문장의 번역을 붙임 (기본 판별력 확인)
    NUM_CHANGE  번역문의 수치 변조
    TRUNCATE    번역문 뒷부분을 잘라냄 (누락)
    TERM_SWAP   대상 언어 용어를 다른 용어로 치환 (사전 필요)

    TERM_SWAP 이 핵심이다. 갑상선→식도 유형을 교차 조건에서 잡는지 본다.

사용법:
    cd ~/이윤우 && source .venv/bin/activate
    python crosslingual_nli.py --inspect       # 말뭉치 구조만 확인
    python crosslingual_nli.py
    python crosslingual_nli.py --langs en zh --n 400
"""

from __future__ import annotations

import argparse
import csv
import glob
import json
import os
import random
import re
import sys
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Sequence, Tuple

MODEL_ID = "MoritzLaurer/mDeBERTa-v3-base-xnli-multilingual-nli-2mil7"
DATA_ROOT = "/home/hufs/shared/data"
CORPUS_DIRS = {"en": "corpus_enko", "zh": "corpus_zhko", "ja": "corpus_jako"}
LANG_NAMES = {"en": "영어", "zh": "중국어", "ja": "일본어", "vi": "베트남어"}

KINDS = ("MISMATCH", "NUM_CHANGE", "TRUNCATE", "TERM_SWAP")


# ===========================================================================
# 유틸
# ===========================================================================

def clean(t: str) -> str:
    t = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", " ", str(t))
    t = unicodedata.normalize("NFC", t)
    return re.sub(r"\s+", " ", t).strip()


def hangul_ratio(t: str) -> float:
    if not t:
        return 0.0
    ko = sum(1 for c in t if "가" <= c <= "힣")
    letters = sum(1 for c in t if c.isalpha() or "\u4e00" <= c <= "\u9fff")
    return ko / max(letters, 1)


def auroc(pos: Sequence[float], neg: Sequence[float]) -> float:
    """pos(훼손)의 점수가 낮을수록 좋음. 1.0 = 완전 분리."""
    if not pos or not neg:
        return float("nan")
    p, n = [-x for x in pos], [-x for x in neg]
    allv = sorted(p + n)
    ranks, i = {}, 0
    while i < len(allv):
        j = i
        while j + 1 < len(allv) and allv[j + 1] == allv[i]:
            j += 1
        ranks[allv[i]] = (i + j) / 2.0 + 1.0
        i = j + 1
    r = sum(ranks[x] for x in p)
    return (r - len(p) * (len(p) + 1) / 2.0) / (len(p) * len(n))


def wilson(k: int, n: int, z: float = 1.96) -> Tuple[float, float]:
    if n == 0:
        return (float("nan"), float("nan"))
    p = k / n
    d = 1 + z * z / n
    c = p + z * z / (2 * n)
    m = z * ((p * (1 - p) / n + z * z / (4 * n * n)) ** 0.5)
    return ((c - m) / d, (c + m) / d)


def tau_for_spec(scores: Sequence[float], target: float) -> float:
    xs = sorted(scores)
    k = max(0, min(int(round((1.0 - target) * len(xs))), len(xs) - 1))
    return xs[k]


# ===========================================================================
# 말뭉치 로딩 (형식 자동 감지)
# ===========================================================================

class CorpusLoader:
    """json / jsonl / csv / tsv / xlsx 를 훑어 한국어-외국어 열을 자동 판별한다."""

    def __init__(self, root: str, verbose: bool = True):
        self.root = root
        self.verbose = verbose

    def files(self, subdir: str) -> List[str]:
        base = os.path.join(self.root, subdir)
        if not os.path.isdir(base):
            return []
        out = []
        for ext in ("json", "jsonl", "csv", "tsv", "xlsx", "xls"):
            out.extend(glob.glob(os.path.join(base, "**", f"*.{ext}"),
                                 recursive=True))
        return sorted(out)

    # -- 개별 파일에서 레코드 목록 뽑기 --------------------------------------

    def _records(self, path: str, limit: int) -> List[dict]:
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == ".jsonl":
                out = []
                with open(path, encoding="utf-8", errors="replace") as f:
                    for line in f:
                        if len(out) >= limit:
                            break
                        line = line.strip()
                        if line:
                            out.append(json.loads(line))
                return out
            if ext == ".json":
                with open(path, encoding="utf-8", errors="replace") as f:
                    data = json.load(f)
                return self._flatten(data, limit)
            if ext in (".csv", ".tsv"):
                delim = "\t" if ext == ".tsv" else ","
                with open(path, encoding="utf-8-sig", errors="replace",
                          newline="") as f:
                    return [r for _, r in zip(range(limit), csv.DictReader(f, delimiter=delim))]
            if ext in (".xlsx", ".xls"):
                try:
                    import openpyxl
                except ImportError:
                    return []
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = ws.iter_rows(values_only=True)
                header = [str(h) if h is not None else f"c{i}"
                          for i, h in enumerate(next(rows, []))]
                out = []
                for _, r in zip(range(limit), rows):
                    out.append({h: v for h, v in zip(header, r)})
                wb.close()
                return out
        except Exception as e:
            if self.verbose:
                print(f"    (읽기 실패 {os.path.basename(path)}: {type(e).__name__})")
        return []

    @staticmethod
    def _flatten(data, limit: int) -> List[dict]:
        """최상위가 dict 인 경우 리스트를 품은 키를 찾아 들어간다."""
        if isinstance(data, list):
            return [d for d in data[:limit] if isinstance(d, dict)]
        if isinstance(data, dict):
            for v in data.values():
                if isinstance(v, list) and v and isinstance(v[0], dict):
                    return v[:limit]
        return []

    # -- 한국어 / 외국어 열 판별 --------------------------------------------

    @staticmethod
    def _pick_columns(recs: List[dict]) -> Optional[Tuple[str, str]]:
        if not recs:
            return None
        keys = [k for k in recs[0] if isinstance(recs[0].get(k), str)
                or recs[0].get(k) is None]
        stats = {}
        for k in keys:
            vals = [str(r.get(k, "")) for r in recs[:200] if r.get(k)]
            vals = [v for v in vals if len(v) >= 8]
            if len(vals) < 5:
                continue
            stats[k] = (sum(hangul_ratio(v) for v in vals) / len(vals),
                        sum(len(v) for v in vals) / len(vals))
        if not stats:
            return None
        ko = [k for k, (h, _) in stats.items() if h >= 0.4]
        fo = [k for k, (h, _) in stats.items() if h < 0.1]
        if not ko or not fo:
            return None
        ko.sort(key=lambda k: -stats[k][1])
        fo.sort(key=lambda k: -stats[k][1])
        return (ko[0], fo[0])

    def load(self, lang: str, n: int, inspect: bool = False) -> List[Tuple[str, str]]:
        sub = CORPUS_DIRS.get(lang)
        if not sub:
            return []
        files = self.files(sub)
        if self.verbose:
            print(f"  [{LANG_NAMES[lang]}] {sub}/ 파일 {len(files)}개")
        pairs: List[Tuple[str, str]] = []
        for path in files:
            recs = self._records(path, 4000)
            cols = self._pick_columns(recs)
            if not cols:
                continue
            kc, fc = cols
            if self.verbose:
                print(f"    {os.path.basename(path)[:50]:<52} "
                      f"열: {kc} / {fc}  ({len(recs)}행)")
            if inspect:
                for r in recs[:2]:
                    print(f"       KO: {clean(r.get(kc,''))[:70]}")
                    print(f"       FO: {clean(r.get(fc,''))[:70]}")
                continue
            for r in recs:
                ko, fo = clean(r.get(kc, "")), clean(r.get(fc, ""))
                if 15 <= len(ko) <= 250 and 10 <= len(fo) <= 400 \
                        and hangul_ratio(ko) >= 0.4 and hangul_ratio(fo) < 0.1:
                    pairs.append((ko, fo))
            if len(pairs) >= n * 8:
                break
        return pairs


# ===========================================================================
# 훼손 주입
# ===========================================================================

class Corrupter:
    NUM = re.compile(r"(?<![\d.])(\d{1,4})")

    def __init__(self, lang: str, termdict: Dict[str, str], rng: random.Random):
        self.lang = lang
        self.rng = rng
        # 대상 언어 용어 목록 (값 쪽). 서로 바꿔치기할 재료로 쓴다.
        self.terms = sorted({v for v in termdict.values() if v and len(str(v)) >= 2})

    def mismatch(self, fo: str, others: Sequence[str]) -> Optional[str]:
        for _ in range(10):
            cand = self.rng.choice(others)
            if cand != fo:
                return cand
        return None

    def num_change(self, fo: str) -> Optional[str]:
        m = self.NUM.search(fo)
        if not m:
            return None
        old = int(m.group(1))
        return fo[:m.start()] + str(old * 3 + 1) + fo[m.end():]

    def truncate(self, fo: str) -> Optional[str]:
        if len(fo) < 40:
            return None
        cut = fo[: int(len(fo) * 0.55)].rstrip()
        return cut if len(cut) >= 15 else None

    def term_swap(self, fo: str) -> Optional[Tuple[str, str]]:
        if not self.terms:
            return None
        present = [t for t in self.terms if t.lower() in fo.lower()]
        if not present:
            return None
        src = max(present, key=len)
        alts = [t for t in self.terms
                if t != src and t.lower() not in fo.lower() and len(t) >= 2]
        if not alts:
            return None
        dst = self.rng.choice(alts)
        i = fo.lower().find(src.lower())
        return (fo[:i] + dst + fo[i + len(src):], f"{src}→{dst}")


# ===========================================================================
# 채점
# ===========================================================================

class Scorer:
    def __init__(self, batch_size: int = 64, max_length: int = 256):
        import torch
        from transformers import AutoModelForSequenceClassification, AutoTokenizer
        self.torch = torch
        self.bs, self.ml = batch_size, max_length
        self.device = "cuda" if torch.cuda.is_available() else "cpu"
        print(f"[INIT] device={self.device} model={MODEL_ID}")
        self.tok = AutoTokenizer.from_pretrained(MODEL_ID)
        self.model = AutoModelForSequenceClassification.from_pretrained(MODEL_ID)
        self.model.to(self.device).eval()
        id2 = {int(k): str(v).lower() for k, v in self.model.config.id2label.items()}
        self.i_ent = next(i for i, v in id2.items() if v.startswith("entail"))
        print(f"[INIT] id2label={id2} -> entail={self.i_ent}")

    def entail(self, premises: List[str], hypotheses: List[str]) -> List[float]:
        torch = self.torch
        out: List[float] = []
        for i in range(0, len(premises), self.bs):
            enc = self.tok(premises[i:i + self.bs], hypotheses[i:i + self.bs],
                           truncation=True, padding=True, max_length=self.ml,
                           return_tensors="pt")
            enc = {k: v.to(self.device) for k, v in enc.items()}
            with torch.no_grad():
                p = torch.softmax(self.model(**enc).logits, dim=-1)
            out.extend(p[:, self.i_ent].tolist())
            if (i // self.bs) % 20 == 0 and i:
                print(f"    {i}/{len(premises)}", flush=True)
        return out


# ===========================================================================

@dataclass
class Row:
    lang: str
    kind: str
    detail: str
    ko: str
    fo: str
    score: float = 0.0


def load_termdict(path: str, lang: str) -> Dict[str, str]:
    """{한국어: 대상언어} 형태로 최대한 유연하게 읽는다."""
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
    except Exception:
        return {}
    out: Dict[str, str] = {}
    for k, v in (d.items() if isinstance(d, dict) else []):
        if isinstance(v, str):
            out[k] = v
        elif isinstance(v, dict):
            for cand in (lang, LANG_NAMES.get(lang, ""), f"{lang}_term", "term"):
                if cand and isinstance(v.get(cand), str):
                    out[k] = v[cand]
                    break
        elif isinstance(v, list) and v and isinstance(v[0], str):
            out[k] = v[0]
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=DATA_ROOT)
    ap.add_argument("--langs", nargs="+", default=["en", "zh", "ja"])
    ap.add_argument("--n", type=int, default=300, help="언어당 정답쌍 표본 수")
    ap.add_argument("--termdict", default="termdict_verified.json")
    ap.add_argument("--target-spec", type=float, default=0.95)
    ap.add_argument("--seed", type=int, default=20260827)
    ap.add_argument("--inspect", action="store_true", help="말뭉치 구조만 출력")
    ap.add_argument("--out", default="crosslingual_nli")
    a = ap.parse_args()

    rng = random.Random(a.seed)
    loader = CorpusLoader(a.root)

    print("[LOAD] 병렬 말뭉치 탐색")
    data: Dict[str, List[Tuple[str, str]]] = {}
    for lang in a.langs:
        pairs = loader.load(lang, a.n, inspect=a.inspect)
        if a.inspect:
            continue
        rng.shuffle(pairs)
        data[lang] = pairs[: a.n]
        print(f"    → 사용 가능 쌍 {len(data[lang])}건")
    if a.inspect:
        print("\n구조 확인용 실행이었습니다. 열 판별이 맞으면 --inspect 없이 실행하세요.")
        return
    if not any(data.values()):
        sys.exit("[FATAL] 사용 가능한 병렬쌍이 없습니다. --inspect 로 구조를 확인하세요.")

    scorer = Scorer()
    rows: List[Row] = []

    for lang, pairs in data.items():
        if not pairs:
            continue
        td = load_termdict(a.termdict, lang)
        print(f"\n[{LANG_NAMES[lang]}] 정답쌍 {len(pairs)}건 / "
              f"용어사전 {len(td)}개")
        corr = Corrupter(lang, td, rng)
        others = [fo for _, fo in pairs]

        batch: List[Row] = []
        for ko, fo in pairs:
            batch.append(Row(lang, "CORRECT", "-", ko, fo))
            m = corr.mismatch(fo, others)
            if m:
                batch.append(Row(lang, "MISMATCH", "-", ko, m))
            nc = corr.num_change(fo)
            if nc:
                batch.append(Row(lang, "NUM_CHANGE", "-", ko, nc))
            tr = corr.truncate(fo)
            if tr:
                batch.append(Row(lang, "TRUNCATE", "-", ko, tr))
            ts = corr.term_swap(fo)
            if ts:
                batch.append(Row(lang, "TERM_SWAP", ts[1], ko, ts[0]))

        print(f"  채점 대상 {len(batch)}건")
        scores = scorer.entail([r.ko for r in batch], [r.fo for r in batch])
        for r, s in zip(batch, scores):
            r.score = s
        rows.extend(batch)

    # -- 집계 ---------------------------------------------------------------
    print("\n" + "=" * 76)
    print(f"언어 교차 NLI 결과   (정상 통과율 {a.target_spec:.0%} 로 통일)")
    print("=" * 76)
    summary = {}
    for lang in data:
        sel = [r for r in rows if r.lang == lang]
        if not sel:
            continue
        pos = [r for r in sel if r.kind == "CORRECT"]
        if not pos:
            continue
        tau = tau_for_spec([r.score for r in pos], a.target_spec)
        print(f"\n[{LANG_NAMES[lang]}]  정답쌍 {len(pos)}건  "
              f"평균 entailment {sum(r.score for r in pos)/len(pos):.3f}  "
              f"보정 임계값 {tau:.3f}")
        print(f"  {'훼손 유형':<14}{'n':>5}{'평균점수':>10}{'탐지율':>10}{'95% CI':>18}")
        summary[lang] = {"tau": round(tau, 4), "n_correct": len(pos), "by_kind": {}}
        for kind in KINDS:
            k_rows = [r for r in sel if r.kind == kind]
            if not k_rows:
                print(f"  {kind:<14}{0:>5}{'—':>10}{'—':>10}")
                continue
            det = sum(1 for r in k_rows if r.score < tau)
            mean = sum(r.score for r in k_rows) / len(k_rows)
            lo, hi = wilson(det, len(k_rows))
            a_ = auroc([r.score for r in k_rows], [r.score for r in pos])
            summary[lang]["by_kind"][kind] = {
                "n": len(k_rows), "mean": round(mean, 4),
                "detect": round(det / len(k_rows), 4), "auroc": round(a_, 4)}
            print(f"  {kind:<14}{len(k_rows):>5}{mean:>10.3f}"
                  f"{det/len(k_rows):>9.1%}   [{lo:.0%},{hi:.0%}]"
                  f"   AUROC {a_:.3f}")

    # -- 판정 ---------------------------------------------------------------
    print("\n" + "=" * 76)
    print("판정")
    print("=" * 76)
    for lang, s in summary.items():
        mm = s["by_kind"].get("MISMATCH", {}).get("auroc")
        ts = s["by_kind"].get("TERM_SWAP", {}).get("auroc")
        print(f"\n[{LANG_NAMES[lang]}]")
        if mm is None:
            print("  MISMATCH 표본 없음 — 판정 불가")
            continue
        if mm < 0.80:
            print(f"  MISMATCH AUROC {mm:.3f} — 전혀 다른 문장조차 구분하지 못합니다.")
            print("  → 이 언어쌍에서 교차 NLI 는 성립하지 않습니다. ⑥ 통합 불가.")
        else:
            print(f"  MISMATCH AUROC {mm:.3f} — 기본 판별력 확인됨.")
            if ts is None:
                print("  TERM_SWAP 표본 없음 (용어사전 매칭 실패) — 별도 확인 필요")
            elif ts < 0.75:
                print(f"  TERM_SWAP AUROC {ts:.3f} — 용어 오역은 여전히 못 잡습니다.")
                print("  → 예상대로입니다. 교차 NLI 는 수치·누락 담당, 용어는 사전 대조 담당.")
            else:
                print(f"  TERM_SWAP AUROC {ts:.3f} — 용어 오역까지 잡습니다.")
                print("  → 사전 의존도를 낮출 수 있습니다. 표본을 늘려 재확인 권장.")
    print("=" * 76)

    with open(f"{a.out}_report.json", "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    with open(f"{a.out}_cases.csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["lang", "kind", "detail", "score", "ko", "fo"])
        for r in rows:
            w.writerow([r.lang, r.kind, r.detail, round(r.score, 4),
                        r.ko[:300], r.fo[:300]])
    print(f"\n[SAVE] {a.out}_report.json / {a.out}_cases.csv")


if __name__ == "__main__":
    main()
